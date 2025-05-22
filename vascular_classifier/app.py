#Importing the necessary libraries.
import io
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import (accuracy_score, f1_score, precision_score, recall_score, roc_auc_score, roc_curve)
from sklearn.preprocessing import label_binarize
from vascular_classifier.model_utils import predict_batch, model
from openpyxl.drawing.image import Image as XLImage  
from openpyxl.utils import get_column_letter  

#Setting the page configuration and hiding the default UI.
st.set_page_config(page_title = "Vascular Condition Classifier", page_icon = "🫀",layout = "wide")
st.markdown(""" <style> #MainMenu {visibility: hidden;} footer {visibility: hidden;} .block-container {padding-top: 2rem;} h1 {color: #E91E63; font-size: 2.8rem;} </style> """, unsafe_allow_html = True)

#Adding the title and caption.
st.title("🫀 Vascular Condition Classifier 🫀")
st.caption("Upload vascular image metadata and get predictions with confidence scores.")

#Ading a sidebar for file upload.
with st.sidebar:
    st.header("Upload Patient Metadata")
    uploaded_file = st.file_uploader("Choose a CSV file", type="csv")

#Adding the logic for when a file is uploaded.
if uploaded_file:
    #Loading and normalizing the CSV.
    df = pd.read_csv(uploaded_file)
    df.columns = df.columns.str.replace('<','_', regex = False).str.replace('>','_', regex = False)
    orig = "Vascular class(0=no, 1=Hyalinosis, 2=Fibrosis, 3=Hyalinosis+Fibrosis)"
    if orig in df.columns:
        df = df.rename(columns = {orig: "Actual Class"})
        df["Actual Class"] = df["Actual Class"].astype(int)
    
    #Creating the preview and prediction tabs.
    tab1, tab2 = st.tabs(["📄 Preview Data", "📈 Predictions and Metrics"])
    with tab1:
        st.dataframe(df.head())
        
    #The predictions tab.
    with tab2:
        if st.button("Run Predictions"):
            with st.spinner("Running..."):
                #Generating the raw predictions and probabilities.
                raw_results = predict_batch(df)
                proba_cols = [f"Prob_Class_{c}" for c in model.classes_]
                y_proba = raw_results[proba_cols].values
                y_pred = raw_results["Predicted Class"].values

                #Creating and display results with rounded values.
                results = raw_results.copy()
                for c in results.select_dtypes("float"):
                    results[c] = results[c].round(3)

                #Attaching the actual class if present.
                if "Actual Class" in df.columns:
                    results["Actual Class"] = df["Actual Class"].values

                    #Computing the performance metrics.
                    y_true = df["Actual Class"]
                    acc   = accuracy_score(y_true, y_pred)
                    f1_w  = f1_score(y_true, y_pred, average = "weighted")
                    f1_m  = f1_score(y_true, y_pred, average = "macro")
                    prec_w= precision_score(y_true, y_pred, average = "weighted")
                    prec_m= precision_score(y_true, y_pred, average = "macro")
                    rec_w = recall_score(y_true, y_pred, average = "weighted")
                    rec_m = recall_score(y_true, y_pred, average = "macro")

                    unique_classes = np.unique(y_true)
                    if len(unique_classes) > 1:
                        auc_w = roc_auc_score(y_true, y_proba, multi_class = "ovr", average = "weighted")
                        auc_m = roc_auc_score(y_true, y_proba, multi_class = "ovr", average = "macro")
                    else:
                        auc_w = auc_m = None

                    #Displaying the metrics as a markdown.
                    metrics_md = (
                        f"- **Accuracy:** {acc:.3f}  \n"
                        f"- **F1 (Weighted / Macro):** {f1_w:.3f} / {f1_m:.3f}  \n"
                        f"- **Precision (Weighted / Macro):** {prec_w:.3f} / {prec_m:.3f}  \n"
                        f"- **Recall (Weighted / Macro):** {rec_w:.3f} / {rec_m:.3f}"
                    )
                    if auc_w is not None:
                        metrics_md += f"  \n- **AUC (Weighted / Macro):** {auc_w:.3f} / {auc_m:.3f}"
                    st.subheader("Performance Metrics")
                    st.markdown(metrics_md)

                    #Plotting the ROC curves for each class.
                    fig = None
                    if len(unique_classes) > 1:
                        y_true_bin = label_binarize(y_true, classes = model.classes_)
                        label_map = {
                            0: "No condition",
                            1: "Hyalinosis",
                            2: "Fibrosis",
                            3: "Hyalinosis + Fibrosis"}
                        fig, ax = plt.subplots(figsize = (4, 3))
                        for idx, cls in enumerate(model.classes_):
                            if cls in unique_classes:
                                fpr, tpr, _ = roc_curve(y_true_bin[:, idx], y_proba[:, idx])
                                ax.plot(fpr, tpr, label = label_map.get(cls, f"Class {cls}"))
                        ax.plot([0,1], [0,1], 'k--', alpha=0.3)
                        ax.set_title('ROC Curves by Class')
                        ax.set_xlabel('False Positive Rate')
                        ax.set_ylabel('True Positive Rate')
                        ax.legend(fontsize = 'small')
                        st.pyplot(fig)
                    else:
                        st.warning("ROC AUC and ROC curves require at least two classes in the data.")

                #Define styling functions.
                def _highlight_mismatch(row):
                    return [
                        "background-color: #D32F2F; color: white;"
                        if (col == "Predicted Class" and row["Predicted Class"] != row.get("Actual Class", None))
                        else ""
                        for col in row.index]
                def _highlight_conf(v):
                    if   v >= 0.95: return "background-color: #388E3C; color: white;"
                    elif v >= 0.60: return "background-color: #FBC02D; color: black;"
                    else:           return "background-color: #D32F2F; color: white;"

                #Applying the styling.
                styler = results.style.format(precision = 2)
                if "Actual Class" in results.columns:
                    styler = styler.apply(_highlight_mismatch, axis = 1)
                if "Confidence Score" in results.columns:
                    styler = styler.applymap(_highlight_conf, subset = ["Confidence Score"])

                st.subheader("Predictions:")
                st.dataframe(styler, use_container_width=True)

                #Exporting the styled results to Excel, with the metrics and chart positioned two columns right of data for cleanliness.
                towrite = io.BytesIO()
                with pd.ExcelWriter(towrite, engine = "openpyxl") as writer:
                    #Writing the stylized table.
                    styler.to_excel(writer, sheet_name = "Predictions", index = False)
                    ws = writer.sheets["Predictions"]
                    #Determining the placement of the data in the Excel sheet.
                    max_col = ws.max_column
                    start_col = max_col + 2
                    #Initializing the display of the metrics.
                    meta = [("Accuracy", acc), ("F1 (w/m)", f"{f1_w:.3f} / {f1_m:.3f}"), ("Precision (w/m)", f"{prec_w:.3f} / {prec_m:.3f}"),("Recall (w/m)", f"{rec_w:.3f} / {rec_m:.3f}"),]
                    if auc_w is not None:
                        meta.append(("AUC (w/m)", f"{auc_w:.3f} / {auc_m:.3f}"))
                    for i, (name, val) in enumerate(meta, start = 1):
                        cell = ws.cell(row = i, column = start_col)
                        cell.value = f"{name}: {val}"

                    #Embedding the created ROC curve in the Excel sheet.
                    if fig is not None:
                        img_stream = io.BytesIO()
                        fig.savefig(img_stream, format = 'png', bbox_inches = 'tight')
                        img_stream.seek(0)
                        img = XLImage(img_stream)
                        
                        #Anchoring the image two columns right and below the metrics.
                        anchor_cell = get_column_letter(start_col) + str(len(meta) + 2)
                        img.anchor = anchor_cell
                        ws.add_image(img)
                towrite.seek(0)

                #Allowing for the download of the Excel sheet.
                st.download_button(
                    "📥 Download Predictions (Excel)",
                    data = towrite,
                    file_name = "vascular_predictions.xlsx",
                    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")